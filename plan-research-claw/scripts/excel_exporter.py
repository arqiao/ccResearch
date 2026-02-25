"""Excel导出 + Web数据生成模块"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("序号", 6), ("标题", 40), ("链接", 30), ("日期", 12), ("来源", 20),
    ("一级分类", 12), ("二级分类", 14), ("内容深度", 12),
    ("摘要", 50), ("字数", 8), ("抓取状态", 10), ("record_id", 16),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def load_articles(enriched_path: str) -> list:
    with open(enriched_path, "r", encoding="utf-8") as f:
        return json.load(f)


def export_excel(config: dict, root_dir: str):
    """导出Excel文件"""
    enriched_path = os.path.join(root_dir, config["paths"]["enriched_json"])
    excel_path = os.path.join(root_dir, config["paths"]["excel_path"])
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    articles = load_articles(enriched_path)
    articles.sort(key=lambda x: x.get("date", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "OpenClaw知识库"

    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, art in enumerate(articles, 2):
        values = [
            row_idx - 1, art.get("title", ""), art.get("url", ""),
            art.get("date", ""), art.get("source", ""),
            art.get("category", ""), art.get("sub_category", ""),
            art.get("depth", ""), art.get("summary", "")[:200],
            art.get("word_count", 0), art.get("scrape_status", ""),
            art.get("record_id", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

        url = art.get("url", "")
        if url:
            ws.cell(row=row_idx, column=2).hyperlink = url
            ws.cell(row=row_idx, column=2).font = Font(color="0563C1", underline="single")

    wb.save(excel_path)
    print(f"✓ Excel已导出: {excel_path} ({len(articles)} 条)")


def generate_web_data(config: dict, root_dir: str):
    """生成 site/articles/ 的数据文件"""
    enriched_path = os.path.join(root_dir, config["paths"]["enriched_json"])
    cat_json_path = os.path.join(root_dir, "data", "categories.json")
    out_dir = os.path.join(root_dir, config["paths"]["site_articles"])
    os.makedirs(out_dir, exist_ok=True)

    articles = load_articles(enriched_path)
    articles.sort(key=lambda x: x.get("date", ""), reverse=True)

    web_data = []
    for art in articles:
        web_data.append({
            "title": art.get("title", ""),
            "url": art.get("url", ""),
            "date": art.get("date", ""),
            "source": art.get("source", ""),
            "category": art.get("category", ""),
            "sub_category": art.get("sub_category", ""),
            "depth": art.get("depth", ""),
            "summary": art.get("summary", "")[:300],
            "word_count": art.get("word_count", 0),
            "hotness": art.get("hotness", 0),
        })

    # 读取分类配置
    with open(cat_json_path, "r", encoding="utf-8") as f:
        categories = json.load(f)

    # 文章数据
    js_path = os.path.join(out_dir, "articles_data.js")
    js_content = "const ARTICLES_DATA = " + json.dumps(web_data, ensure_ascii=False, indent=2) + ";\n"
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js_content)

    # 分类配置
    cat_js_path = os.path.join(out_dir, "categories.js")
    cat_js = "const CATEGORIES = " + json.dumps(categories, ensure_ascii=False, indent=2) + ";\n"
    with open(cat_js_path, "w", encoding="utf-8") as f:
        f.write(cat_js)

    print(f"✓ Web数据已生成: {out_dir} ({len(web_data)} 条)", flush=True)
